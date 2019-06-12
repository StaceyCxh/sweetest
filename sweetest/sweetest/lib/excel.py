import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors
from openpyxl.styles.fills import PatternFill
from sweetest.globals import g


DarkRedFill = PatternFill("solid", fgColor="CD3700")
RedFill = PatternFill("solid", fgColor="EE0000")
GreenFill = PatternFill("solid", fgColor="006400")
OrangeFill = PatternFill("solid", fgColor="EE4000")
BlueFill = PatternFill("solid", fgColor="0000FF")
GrayFill = PatternFill("solid", fgColor="7A7A7A")

Fills = {
    'Fail': RedFill,
    'Pass': GreenFill,
    'Skip': BlueFill,
    'Block': OrangeFill,
}


class Excel(object):
    '''
    Excel类，用于处理excel
    '''
    def __init__(self, file_name):
        '''
        初始化excel文件，可读写
        :param file_name: excel文件名
        '''
        super(Excel, self).__init__()
        self.length = 0
        self.workbook = load_workbook(file_name)

    def get_sheet(self, sheet_name):
        '''
        匹配获取excel的标签页名
        :param sheet_name: 标签页名，可模糊匹配
        :return: 匹配的标签页名
        '''
        names = []
        flag = 0b00
        if isinstance(sheet_name, str):
            if sheet_name.startswith('^'):
                flag = flag | 0b01
                sheet_name = sheet_name[1:]
            elif sheet_name.startswith('*'):
                flag = flag | 0b100
                sheet_name = sheet_name[1:]

            if sheet_name.endswith('$'):
                flag = flag | 0b10
                sheet_name = sheet_name[:-1]
            elif sheet_name.endswith('*'):
                flag = flag | 0b100
                sheet_name = sheet_name[:-1]

            if flag == 0b00 or (flag & 0b11 == 0b11):
                for name in self.workbook.get_sheet_names():
                    if sheet_name == name:
                        names.append(name)
                        break
            elif flag & 0b01:
                for name in self.workbook.get_sheet_names():
                    if name.startswith(sheet_name):
                        names.append(name)
            elif flag & 0b10:
                for name in self.workbook.get_sheet_names():
                    if name.endswith(sheet_name):
                        names.append(name)
            else:
                for name in self.workbook.get_sheet_names():
                    if sheet_name in name:
                        names.append(name)

        elif isinstance(sheet_name, list):
            names = sheet_name
        else:
            raise Exception('错误: 无效的excel表单名: %s' % sheet_name)

        return names

    def read(self, sheet_name):
        '''
        获取excle表单中的数据
        :param sheet_name: Excel 中标签页名称
        :return: 表单中的数据，格式为二维list，[[],[]……]
        '''

        sheet = self.workbook.get_sheet_by_name(sheet_name)
        data = []
        for row in sheet.rows:
            d = []
            for cell in row:
                if cell.value is None:
                    value = ''
                else:
                    value = cell.value
                d.append(value)
            data.append(d)
        return data

    def write(self):
        '''
        将测试结果写入excel报告中，并进行格式化
        '''
        sheet = self.workbook.get_sheet_by_name(g.current_sheet_name)
        # 首行用例标题设置红色底色、字体白色
        for cell in sheet[1]:
            cell.fill = DarkRedFill
            cell.font = Font(color=colors.WHITE)
        m = 0
        n = 2
        while m < len(g.testsuite):
            case = g.testsuite[m]
            # setup 和 teardown 不写入相关内容
            if case.get('condition', '') in ['setup', 'SETUP', 'teardown', 'TEARDOWN']:
                m = m + 1
                steps = case.get('steps')
                n = n + len(steps)
                continue

            # 将setup插入到用例中，作为操作步骤第0步
            l = 0
            if case.get('condition', '') in ['', None] and case.get('flag', '') in ['', None, 'Y', 'y']:
                l = self.insert_setup(sheet, case.get('setup', {}), n)

            cell_result = '%s%d' % (chr(111), n)
            cell_id = '%s%d' % (chr(97), n)
            cell_title = '%s%d' % (chr(98), n)

            sheet[cell_result].value = case.get('result', '')
            # 用例结果列设置底色
            if sheet[cell_result].value:
                sheet[cell_result].fill = Fills[sheet[cell_result].value]
                sheet[cell_result].font = Font(color=colors.WHITE)

            sheet[cell_id].value = case.get('id', '')
            sheet[cell_title].value = case.get('title', '')

            n = n + l
            if l:
                cell_id = '%s%d' % (chr(97), n)
                cell_title = '%s%d' % (chr(98), n)
                sheet[cell_id].value = None
                sheet[cell_title].value = None

            # 将测试用例各步骤的执行结果写入
            self.length = 0
            for step in case.get('steps'):
                cell_score = '%s%d' % (chr(110), (n+self.length))
                cell_remark = '%s%d' % (chr(112), (n+self.length))
                sheet[cell_score].value = step.get('score', '')
                sheet[cell_score].font = Font(color=colors.BLACK)
                if sheet[cell_score].value == 'NO':
                    sheet[cell_score].fill = GrayFill
                    sheet[cell_score].font = Font(color=colors.WHITE)
                sheet[cell_remark].value = step.get('remark', '')
                sheet[cell_remark].font = Font(color=colors.BLACK)

                z = 1
                if step.get('snippets', '') not in ('', []):
                    z = self.insert_snippet(sheet, step.get('element').split('*')[0], step.get('snippets'), (n+self.length+1)) + 1
                self.length = self.length + z

            n = n + self.length
            # 将teardown插入到用例中，作为操作步骤最后一步
            l = 0
            if case.get('condition', '') in ['', None] and case.get('flag', '') in ['', None, 'Y', 'y']:
                l = self.insert_teardown(sheet, case.get('id'), n)

            m = m + 1
            n = n + l

        self.workbook.save(g.report_file)

    def insert_snippet(self, sheet, caseID, rows, no):
        m = 0
        for row in rows:
            n = len(row)
            # 插入空白行
            sheet.insert_rows(no, n)
            # 空白行写入测试片段各步骤的内容
            z = self.insert_rows(sheet, caseID, no, row)
            m = m + z
            no = no + z
        return m

    def insert_setup(self, sheet, case, no):
        return self.insert_fixture(sheet, 0, case, no)

    def insert_teardown(self, sheet, caseID, no):
        for i, case in enumerate(g.normal_testcases):
            if case.get('id') == caseID and g.teardowns:
                case = g.teardowns[i]
                pre_no = case.get('steps')[-1].get('no')
                break
        if pre_no[0] in ['^', '<', '>']:
            pre_no = pre_no[1:]

        return self.insert_fixture(sheet, 1, case, no, (int(pre_no)+1))

    def insert_fixture(self, sheet, flag, case, no, pre_no=0):
        sheet.insert_rows(no, (len(case.get('steps')) + 1))

        cell_no = '%s%d' % (chr(100), no)
        cell_keyword = '%s%d' % (chr(101), no)
        cell_page = '%s%d' % (chr(102), no)
        cell_element = '%s%d' % (chr(103), no)
        cell_score = '%s%d' % (chr(110), no)
        if flag == 0:
            sheet[cell_page].value = 'SETUP'
        else:
            sheet[cell_page].value = 'TEARDOWN'
        sheet[cell_no].value = pre_no
        sheet[cell_keyword].value = '执行'
        sheet[cell_element].value = str(case.get('id', ''))
        if case.get('result', '') == 'Pass':
            sheet[cell_score].value = 'OK'
        else:
            sheet[cell_score].value = 'NO'
        no = no + 1

        row = case.get('steps', [])
        caseID = str(case.get('id', ''))
        m = self.insert_rows(sheet, caseID, no, row)
        return m + 1

    def insert_rows(self, sheet, caseID, no, row):
        m = 0
        for step in row:
            cell_no = '%s%d' % (chr(100), no)
            cell_keyword = '%s%d' % (chr(101), no)
            cell_page = '%s%d' % (chr(102), no)
            cell_element = '%s%d' % (chr(103), no)
            cell_data = '%s%d' % (chr(104), no)
            cell_expected = '%s%d' % (chr(105), no)
            cell_output = '%s%d' % (chr(106), no)
            cell_score = '%s%d' % (chr(110), no)
            cell_remark = '%s%d' % (chr(112), no)
            sheet[cell_no].value = step.get('no', '')
            sheet[cell_keyword].value = step.get('keyword', '')
            sheet[cell_page].value = step.get('page', '')
            sheet[cell_element].value = str(step.get('element', ''))
            sheet[cell_data].value = str(step.get('data', '')) if step.get('data', '') else None
            sheet[cell_expected].value = str(step.get('expected')) if step.get('expected', '') else None
            sheet[cell_output].value = str(step.get('output')) if step.get('output', '') else None
            sheet[cell_score].value = step.get('score', '')
            sheet[cell_score].font = Font(color=colors.BLACK)
            if sheet[cell_score].value == 'NO':
                sheet[cell_score].fill = GrayFill
                sheet[cell_score].font = Font(color=colors.WHITE)
            sheet[cell_remark].value = step.get('remark', '')
            cell_title = '%s%d' % (chr(99), no)
            sheet[cell_title].value = caseID
            z = 1
            if step.get('snippets', '') not in ('', []):
                z = self.insert_snippet(sheet, step.get('element').split('*')[0], step.get('snippets'), (no + 1)) + 1
            m = m + z
            no = no + z
        return m

    @classmethod
    def copy_excel(cls, filename1, filename2):
        '''
        复制excel文件
        :param filename1: 被复制的文件
        :param filename2: 复制后的文件
        '''
        wb2 = Workbook().save(filename2)
        wb1 = load_workbook(filename1)
        wb2 = load_workbook(filename2)
        sheets1 = wb1.get_sheet_names()
        for i in range(len(sheets1)-1):
            wb2.create_sheet()
        sheets2 = wb2.get_sheet_names()
        for index, sheet1 in enumerate(sheets1):
            sheet1 = wb1.get_sheet_by_name(sheet1)
            sheet2 = wb2.get_sheet_by_name(sheets2[index])
            sheet2.title = sheet1.title
            max_row = sheet1.max_row
            max_column = sheet1.max_column
            for m in range(1, max_row + 1):
                # chr(97) = 'a'
                for n in range(97, 97 + max_column):
                    # 转换成ASCII字符
                    n = chr(n)
                    # 单元格编号
                    i = '%s%d' % (n, m)
                    # 获取文件1中该表单的单元格数据
                    cell1 = sheet1[i].value
                    sheet2[i].value = cell1
        wb2.save(filename2)
        wb1.close()
        wb2.close()

    def close(self):
        '''
        操作完excel后，关闭excel文件
        '''
        self.workbook.close()
